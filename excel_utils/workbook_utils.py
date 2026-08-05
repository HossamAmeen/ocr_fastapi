"""Shared workbook utilities used by all Excel writers."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


def find_sheet(wb: Workbook, name: str) -> Worksheet:
    """Find a sheet by name with whitespace stripping and case-insensitive matching.

    Raises a detailed ValueError showing both requested and available sheet names
    (both raw and normalized) if the sheet is not found.
    """
    target = name.strip().casefold()
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().casefold() == target:
            return wb[sheet_name]

    normalized_available = [s.strip().casefold() for s in wb.sheetnames]
    raise ValueError(
        f"Sheet {name!r} (normalized: {target!r}) not found in workbook.\n"
        f"  Available sheets (raw): {list(wb.sheetnames)}\n"
        f"  Available sheets (normalized): {normalized_available}"
    )


def validate_workbook_sheets(
    workbook_path: str | Path,
    required_sheets: list[str],
) -> None:
    """Startup / pre-flight validation for Excel templates.

    Verifies that the Excel file at workbook_path contains all required sheet names
    (using whitespace-tolerant, case-insensitive comparison).
    Raises ValueError if any required sheet is missing.
    """
    from openpyxl import load_workbook

    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found at {path}")

    wb = load_workbook(path, read_only=True, keep_vba=False)
    try:
        available_normalized = [s.strip().casefold() for s in wb.sheetnames]
        missing = []
        for req in required_sheets:
            target = req.strip().casefold()
            if target not in available_normalized:
                missing.append(req.strip())

        if missing:
            raise ValueError(
                f"Template file '{path.name}' is missing required sheet(s): {missing}\n"
                f"  Available sheets in file: {list(wb.sheetnames)}"
            )
    finally:
        wb.close()
