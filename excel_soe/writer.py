"""Read data.xlsm and append SOE time-log rows from extracted PDF data."""

from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

SOE_SHEET = "SOE"
DATA_START_ROW = 11
FOOTER_START_ROW = 60
TABLE_HEADER_ROW = 10
TABLE_INNER_BORDER_ROW = 12
TABLE_LAST_DATA_BORDER_ROW = 14
TABLE_LAST_COL = 19  # column S
DATE_FORMAT = "dd-mmm-yy"
_FOOTER_MARKER = "CUSTOMER REP"
_FOOTER_BLOCK_ROWS = 9
DEFAULT_TEMPLATE = Path("data.xlsm")
_SHEET_CELL_REF = re.compile(r"^='?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", re.IGNORECASE)


def read_template_rig(workbook_path: str | Path) -> str:
    """Read the Rig value from an SOE workbook.

    Prefers the SOE sheet Rig cell (what the user edits), then falls back to
    Master Data / MS2 formula sources.
    """
    workbook_path = Path(workbook_path)
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)

    def resolve(value: object, depth: int = 0) -> str:
        if value is None or depth > 5:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        if not text.startswith("="):
            return text
        match = _SHEET_CELL_REF.match(text)
        if not match:
            return ""
        sheet_name, col, row = match.group(1), match.group(2), int(match.group(3))
        if sheet_name not in wb.sheetnames:
            return ""
        return resolve(wb[sheet_name][f"{col}{row}"].value, depth + 1)

    def soe_sheet_rig() -> str:
        if "SOE" not in wb.sheetnames:
            return ""
        ws = wb["SOE"]
        # Common layout: N5="Rig:", R5=value (merged R5:S5)
        for coord in ("R5", "O5", "P5", "Q5"):
            value = resolve(ws[coord].value)
            if value:
                return value
        for row in ws.iter_rows(min_row=1, max_row=12, max_col=25):
            for cell in row:
                label = str(cell.value or "").strip().rstrip(":").casefold()
                if label != "rig":
                    continue
                for offset in range(1, 8):
                    neighbor = ws.cell(cell.row, cell.column + offset).value
                    value = resolve(neighbor)
                    if value:
                        return value
        return ""

    # Prefer SOE sheet first — users typically edit Rig there.
    value = soe_sheet_rig()
    if value:
        return value

    if "Master Data" in wb.sheetnames:
        value = resolve(wb["Master Data"]["E11"].value)
        if value:
            return value

    if "MS2" in wb.sheetnames:
        value = resolve(wb["MS2"]["I8"].value)
        if value:
            return value

    # Fallback to cached calculated values when formulas were evaluated in Excel.
    try:
        wb_data = load_workbook(workbook_path, data_only=True, keep_vba=keep_vba)
    except Exception:
        return ""
    for sheet_name, coord in (("SOE", "R5"), ("Master Data", "E11"), ("MS2", "I8")):
        if sheet_name not in wb_data.sheetnames:
            continue
        cached = wb_data[sheet_name][coord].value
        if cached is not None and str(cached).strip():
            return str(cached).strip()
    return ""



def time_log_to_rows(data: dict[str, Any]) -> list[dict[str, str | datetime]]:
    """Convert extract_time_log() output into SOE table rows."""
    period_from = _parse_report_date(data.get("report_period_from", ""))
    period_to = _parse_report_date(data.get("report_period_to", "")) or period_from

    rows: list[dict[str, str | datetime]] = []
    previous_minutes: int | None = None

    for entry in data.get("entries", []):
        end_time = str(entry.get("to") or entry.get("from") or "")
        minutes = _time_to_minutes(end_time)
        event_date = period_from
        if (
            period_from
            and period_to
            and minutes is not None
            and previous_minutes is not None
            and minutes < previous_minutes
        ):
            event_date = period_to
        elif end_time.startswith("0:") and period_to:
            event_date = period_to

        rows.append(
            {
                "date": event_date or "",
                "time": end_time,
                "event": _format_event(entry),
            }
        )
        if minutes is not None:
            previous_minutes = minutes

    oamn_date = period_to or period_from
    for entry in data.get("oamn_entries", []):
        time_from = str(entry.get("from", ""))
        time_to = str(entry.get("to", ""))
        rows.append(
            {
                "date": oamn_date or "",
                "time": f"{time_from} - {time_to}" if time_from and time_to else time_to or time_from,
                "event": _format_oamn_event(entry),
            }
        )

    return rows


def operational_summary_to_rows(data: dict[str, Any]) -> list[dict[str, str | datetime]]:
    """Convert extract_operational_time_summary() output into SOE table rows."""
    report_date = _parse_drilling_date(data.get("date", ""))
    rows: list[dict[str, str | datetime]] = []

    for entry in data.get("entries", []):
        entry_date = _parse_drilling_date(str(entry.get("report_date") or "")) or report_date
        rows.append(
            {
                "date": entry_date or "",
                "time": str(entry.get("from", "")),
                "event": str(entry.get("operation_details", "")).strip(),
            }
        )

    return rows


def soe_data_to_rows(data: dict[str, Any]) -> list[dict[str, str | datetime]]:
    """Convert any supported SOE extraction dict into Excel table rows."""
    if data.get("source") == "operational_time_summary":
        return operational_summary_to_rows(data)
    return time_log_to_rows(data)


def sort_soe_rows(rows: list[dict[str, str | datetime]]) -> list[dict[str, str | datetime]]:
    """Sort SOE rows chronologically by date, then time."""
    return sorted(rows, key=_soe_row_sort_key)


def write_soe_rows(
    input_path: str | Path,
    output_path: str | Path,
    rows: list[dict[str, str | datetime]],
    template_path: str | Path | None = None,
) -> tuple[Path, int]:
    """Write pre-built SOE rows into the workbook table area."""
    input_path = Path(input_path)
    output_path = Path(output_path)
    template_path = Path(template_path or DEFAULT_TEMPLATE)
    if not rows:
        return output_path, 0

    keep_vba = input_path.suffix.lower() == ".xlsm"
    wb = load_workbook(input_path, keep_vba=keep_vba)
    ws = wb[SOE_SHEET]
    template_ws = _load_template_sheet(template_path, keep_vba)

    footer_row = _find_footer_row(ws)
    footer_row = _compact_leading_empty_rows(ws, template_ws, footer_row)
    last_data_row = _find_last_data_row(ws, footer_row)
    write_start = (
        last_data_row + 1
        if last_data_row >= DATA_START_ROW
        else DATA_START_ROW
    )

    available_rows = footer_row - write_start
    if available_rows < len(rows):
        ws.insert_rows(footer_row, amount=len(rows) - available_rows)

    for index, row_data in enumerate(rows):
        _write_table_row(ws, template_ws, write_start + index, row_data)

    last_data_row = write_start + len(rows) - 1
    new_footer_row = last_data_row + 1

    _remove_duplicate_footers(ws, new_footer_row)
    _restore_footer_block(ws, template_ws, new_footer_row)
    _apply_table_borders(ws, template_ws, DATA_START_ROW, last_data_row)
    _update_print_area(ws, new_footer_row + _FOOTER_BLOCK_ROWS - 1)

    wb.save(output_path)
    return output_path, len(rows)


def write_soe_table(
    input_path: str | Path,
    output_path: str | Path,
    data: dict[str, Any],
    template_path: str | Path | None = None,
) -> tuple[Path, int]:
    """Load workbook, expand the data table, then append the footer below it."""
    input_path = Path(input_path)
    output_path = Path(output_path)
    template_path = Path(template_path or DEFAULT_TEMPLATE)
    rows = soe_data_to_rows(data)
    if not rows:
        return output_path, 0

    keep_vba = input_path.suffix.lower() == ".xlsm"
    wb = load_workbook(input_path, keep_vba=keep_vba)
    ws = wb[SOE_SHEET]
    template_ws = _load_template_sheet(template_path, keep_vba)

    footer_row = _find_footer_row(ws)
    footer_row = _compact_leading_empty_rows(ws, template_ws, footer_row)
    last_data_row = _find_last_data_row(ws, footer_row)
    write_start = (
        last_data_row + 1
        if last_data_row >= DATA_START_ROW
        else DATA_START_ROW
    )

    available_rows = footer_row - write_start
    if available_rows < len(rows):
        ws.insert_rows(footer_row, amount=len(rows) - available_rows)

    for index, row_data in enumerate(rows):
        _write_table_row(ws, template_ws, write_start + index, row_data)

    last_data_row = write_start + len(rows) - 1
    new_footer_row = last_data_row + 1

    _remove_duplicate_footers(ws, new_footer_row)
    _restore_footer_block(ws, template_ws, new_footer_row)
    _apply_table_borders(ws, template_ws, DATA_START_ROW, last_data_row)
    _update_print_area(ws, new_footer_row + _FOOTER_BLOCK_ROWS - 1)

    wb.save(output_path)
    return output_path, len(rows)


def _load_template_sheet(template_path: Path, keep_vba: bool) -> Worksheet:
    if not template_path.is_file():
        raise FileNotFoundError(f"SOE template not found: {template_path}")
    template_wb = load_workbook(template_path, keep_vba=keep_vba)
    return template_wb[SOE_SHEET]


def _find_footer_row(ws: Worksheet) -> int:
    for row in range(DATA_START_ROW, ws.max_row + 1):
        value = _cell_value(ws, row, 1)
        if value and str(value).strip().upper().startswith(_FOOTER_MARKER):
            return row
    return FOOTER_START_ROW


def _find_last_data_row(ws: Worksheet, footer_row: int) -> int:
    last_used = DATA_START_ROW - 1
    for row in range(DATA_START_ROW, footer_row):
        if _is_table_data_row(ws, row):
            last_used = row
    return last_used


def _collect_table_entries(ws: Worksheet, footer_row: int) -> list[dict[str, str | datetime]]:
    entries: list[dict[str, str | datetime]] = []
    for row in range(DATA_START_ROW, footer_row):
        if not _is_table_data_row(ws, row):
            continue
        date_value = _cell_value(ws, row, 1) or ""
        entries.append(
            {
                "date": date_value,
                "time": _cell_value(ws, row, 4) or "",
                "event": _cell_value(ws, row, 5) or "",
            }
        )
    return entries


def _compact_leading_empty_rows(
    ws: Worksheet,
    template_ws: Worksheet,
    footer_row: int,
) -> int:
    """Move table data up so the first row after the header is not blank."""
    entries = _collect_table_entries(ws, footer_row)
    if not entries:
        return footer_row

    first_data_row = next(
        row
        for row in range(DATA_START_ROW, footer_row)
        if _is_table_data_row(ws, row)
    )
    if first_data_row <= DATA_START_ROW:
        return footer_row

    for row in range(DATA_START_ROW, footer_row):
        _clear_table_row_layout(ws, row)

    for index, entry in enumerate(entries):
        _write_table_row(ws, template_ws, DATA_START_ROW + index, entry)

    new_footer_row = DATA_START_ROW + len(entries)
    _remove_duplicate_footers(ws, new_footer_row)
    _restore_footer_block(ws, template_ws, new_footer_row)
    return new_footer_row


def _remove_duplicate_footers(ws: Worksheet, from_row: int) -> None:
    row = from_row
    while row <= ws.max_row:
        value = _cell_value(ws, row, 1)
        if value and str(value).strip().upper().startswith(_FOOTER_MARKER):
            _clear_footer_block(ws, row)
            row += _FOOTER_BLOCK_ROWS
            continue
        row += 1


def _apply_table_borders(
    ws: Worksheet,
    template_ws: Worksheet,
    first_row: int,
    last_row: int,
) -> None:
    if last_row < first_row:
        return

    border_cols = (1, 4, 5, TABLE_LAST_COL)
    for row in range(first_row, last_row + 1):
        template_row = (
            TABLE_LAST_DATA_BORDER_ROW
            if row == last_row
            else TABLE_INNER_BORDER_ROW
        )
        for col in border_cols:
            source = template_ws.cell(template_row, col)
            target = ws.cell(row, col)
            if isinstance(source, MergedCell) or isinstance(target, MergedCell):
                top_left = _merged_top_left_col(template_ws, template_row, col)
                if top_left != col:
                    continue
                source = template_ws.cell(template_row, top_left)
                target = ws.cell(row, top_left)
            if isinstance(target, MergedCell):
                continue
            target.border = copy(source.border)


def _update_print_area(ws: Worksheet, last_row: int) -> None:
    ws.print_area = f"A1:S{last_row}"


def _is_table_data_row(ws: Worksheet, row: int) -> bool:
    if not any(_cell_value(ws, row, col) for col in (1, 4, 5)):
        return False
    label = _cell_value(ws, row, 1)
    if label is None:
        return True
    text = str(label).strip().upper()
    return not text.startswith((_FOOTER_MARKER, "SIGNATURE", "©"))


def _write_table_row(
    ws: Worksheet,
    template_ws: Worksheet,
    row: int,
    row_data: dict[str, str | datetime],
) -> None:
    _clear_table_row_layout(ws, row)
    _copy_table_row_layout(ws, template_ws, TABLE_INNER_BORDER_ROW, row)

    date_value = row_data["date"]
    if isinstance(date_value, datetime):
        ws.cell(row, 1).value = date_value
        ws.cell(row, 1).number_format = DATE_FORMAT
    elif date_value:
        ws.cell(row, 1).value = date_value

    ws.cell(row, 4).value = row_data["time"]
    ws.cell(row, 4).number_format = "General"

    event_cell = ws.cell(row, 5)
    event_cell.value = row_data["event"]
    event_cell.alignment = Alignment(
        wrap_text=True,
        vertical="top",
        horizontal=event_cell.alignment.horizontal or "left",
    )


def _clear_table_row_layout(ws: Worksheet, row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= row <= merged.max_row
            and merged.min_col <= TABLE_LAST_COL
            and merged.max_col >= 1
        ):
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                pass

    for col in range(1, TABLE_LAST_COL + 1):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell):
            cell.value = None


def _copy_table_row_layout(
    ws: Worksheet,
    template_ws: Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    source_height = template_ws.row_dimensions[source_row].height
    if source_height:
        ws.row_dimensions[target_row].height = source_height

    for merged in template_ws.merged_cells.ranges:
        if merged.min_row == source_row == merged.max_row and merged.max_col <= TABLE_LAST_COL:
            ws.merge_cells(
                start_row=target_row,
                start_column=merged.min_col,
                end_row=target_row,
                end_column=merged.max_col,
            )

    for col in range(1, TABLE_LAST_COL + 1):
        source = template_ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(source, MergedCell) or isinstance(target, MergedCell):
            top_left_col = _merged_top_left_col(template_ws, source_row, col)
            if top_left_col != col:
                continue
            source = template_ws.cell(source_row, top_left_col)
            target = ws.cell(target_row, top_left_col)
        if isinstance(target, MergedCell):
            continue
        target.number_format = copy(source.number_format)
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.alignment = copy(source.alignment)


def _restore_footer_block(ws: Worksheet, template_ws: Worksheet, footer_row: int) -> None:
    template_footer = _find_footer_row(template_ws)
    _clear_footer_block(ws, footer_row)

    row_offset = footer_row - template_footer
    max_col = max(template_ws.max_column, ws.max_column, TABLE_LAST_COL)

    for merged in template_ws.merged_cells.ranges:
        if merged.min_row >= template_footer and merged.min_row < template_footer + _FOOTER_BLOCK_ROWS:
            ws.merge_cells(
                start_row=merged.min_row + row_offset,
                start_column=merged.min_col,
                end_row=merged.max_row + row_offset,
                end_column=merged.max_col,
            )

    for source_row in range(template_footer, template_footer + _FOOTER_BLOCK_ROWS):
        target_row = source_row + row_offset
        source_height = template_ws.row_dimensions[source_row].height
        if source_height:
            ws.row_dimensions[target_row].height = source_height

        for col in range(1, max_col + 1):
            source = template_ws.cell(source_row, col)
            if isinstance(source, MergedCell):
                top_left_col = _merged_top_left_col(template_ws, source_row, col)
                if top_left_col != col:
                    continue
                source = template_ws.cell(source_row, top_left_col)
                col = top_left_col

            target = ws.cell(target_row, col)
            if isinstance(target, MergedCell):
                for merged in ws.merged_cells.ranges:
                    if target.coordinate in merged:
                        target = ws.cell(merged.min_row, merged.min_col)
                        break
            target.value = source.value
            target.number_format = copy(source.number_format)
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.alignment = copy(source.alignment)


def _clear_footer_block(ws: Worksheet, footer_row: int) -> None:
    end_row = footer_row + _FOOTER_BLOCK_ROWS - 1
    max_col = max(ws.max_column, TABLE_LAST_COL)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= footer_row and merged.min_row <= end_row:
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                pass
    for row in range(footer_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _merged_top_left_col(ws: Worksheet, row: int, col: int) -> int:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return col
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return merged.min_col
    return col


def _format_event(entry: dict[str, Any]) -> str:
    lines = [str(entry.get("operation", "")).strip()]
    for note in entry.get("notes", []):
        note_text = str(note).strip()
        if note_text:
            lines.append(f"* {note_text}")
    return "\n".join(line for line in lines if line)


def _format_oamn_event(entry: dict[str, Any]) -> str:
    lines = [str(entry.get("operation", "")).strip()]
    for note in entry.get("notes", []):
        note_text = str(note).strip()
        if note_text:
            lines.append(f"* {note_text}")
    return "\n".join(line for line in lines if line)


def _soe_row_sort_key(row: dict[str, str | datetime]) -> tuple[datetime, int, str]:
    date_value = row.get("date")
    if isinstance(date_value, datetime):
        parsed_date = date_value
    else:
        parsed_date = _parse_sortable_date(str(date_value or "")) or datetime.min

    time_text = str(row.get("time") or "")
    time_minutes = _time_sort_minutes(time_text)
    return parsed_date, time_minutes, time_text.casefold()


def _parse_sortable_date(value: str) -> datetime | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    if isinstance(value, datetime):
        return value
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    return _parse_report_date(cleaned) or _parse_drilling_date(cleaned)


def _time_sort_minutes(value: str) -> int:
    cleaned = value.strip()
    if not cleaned:
        return -1
    range_match = re.match(r"^(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})$", cleaned)
    if range_match:
        cleaned = range_match.group(1)
    minutes = _time_to_minutes(cleaned)
    return minutes if minutes is not None else -1


def _parse_drilling_date(value: str) -> datetime | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(cleaned, fmt)
        except ValueError:
            continue
    return _parse_report_date(cleaned)


def _parse_report_date(value: str) -> datetime | None:
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", value)
    if not match:
        return None
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return datetime(year, month, day)


def _time_to_minutes(value: str) -> int | None:
    match = re.match(r"^(\d{1,2}):(\d{2})$", value.strip())
    if not match:
        return None
    return int(match.group(1)) * 60 + int(match.group(2))


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col).value
        return None
    return cell.value
