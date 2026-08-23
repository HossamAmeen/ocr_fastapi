"""Read data.xlsm and rewrite the Proforma table from Python data."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from excel_utils.workbook_utils import find_sheet

PROFORMA_SHEET = "Proforma".strip()
HEADER_ROW = 7
DATA_START_ROW = 8
# Row 8 is the template's existing first table row — it is always kept as-is;
# extracted items are appended starting from the row right below it.
APPEND_START_ROW = DATA_START_ROW + 1
GROSS_VALUE_ROW = 15
PROTECTED_HEADER_END_ROW = 7

# Header metadata that must never be overwritten (e.g. DT label + value).
PROTECTED_CELLS = ("D3", "E3", "D4", "E4", "D5", "E5", "D6", "E6")


def write_proforma_table(
    input_path: str | Path,
    output_path: str | Path,
    items: list[dict[str, Any]],
) -> Path:
    """Load xlsm, append Proforma line items below the existing first row, and save.

    Row 8 (the template's existing first table row) is never cleared or
    overwritten; extracted items are appended starting at row 9.
    """
    input_path = Path(input_path)
    output_path = Path(output_path)

    wb = load_workbook(input_path, keep_vba=True)
    ws = find_sheet(wb, PROFORMA_SHEET)

    protected = _snapshot_cells(ws, PROTECTED_CELLS)

    _clear_old_rows(ws, start_row=APPEND_START_ROW, keep_rows=len(items))
    _write_items(ws, items)
    _update_totals(ws, item_count=len(items))
    _restore_cells(ws, protected)

    wb.save(output_path)
    return output_path


def _snapshot_cells(ws: Worksheet, addresses: tuple[str, ...]) -> dict[str, object]:
    return {addr: ws[addr].value for addr in addresses}


def _restore_cells(ws: Worksheet, values: dict[str, object]) -> None:
    for addr, value in values.items():
        cell = ws[addr]
        if isinstance(cell, MergedCell):
            continue
        cell.value = value


def _clear_old_rows(ws: Worksheet, start_row: int, keep_rows: int) -> None:
    """Clear unused data rows before the totals section (never touch header rows)."""
    if start_row <= PROTECTED_HEADER_END_ROW:
        raise ValueError("Cannot clear protected header rows.")

    for row in range(start_row + keep_rows, GROSS_VALUE_ROW):
        for col in ("A", "B", "D", "E", "F"):
            cell = ws[f"{col}{row}"]
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _existing_first_sno(ws: Worksheet) -> int:
    """Return the S/No already in row 8, so appended rows continue counting from it."""
    value = ws[f"A{DATA_START_ROW}"].value
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _write_items(ws: Worksheet, items: list[dict[str, Any]]) -> None:
    template_row = DATA_START_ROW
    base_sno = _existing_first_sno(ws)

    for index, item in enumerate(items):
        row = APPEND_START_ROW + index
        if row >= GROSS_VALUE_ROW:
            break

        _copy_row_style(ws, template_row, row)

        ws[f"A{row}"].value = base_sno + index + 1
        ws[f"B{row}"].value = item["description"]
        ws[f"D{row}"].value = float(item["per_day_rate"])
        ws[f"E{row}"].value = float(item["days"])
        ws[f"F{row}"].value = f"=D{row}*E{row}"


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int) -> None:
    for col in ("A", "B", "D", "E", "F"):
        source = ws[f"{col}{source_row}"]
        target = ws[f"{col}{target_row}"]
        target.number_format = copy(source.number_format)
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.alignment = copy(source.alignment)


def _update_totals(ws: Worksheet, item_count: int) -> None:
    """Sum F8 (the retained first row) through the last appended row."""
    first_row = DATA_START_ROW
    last_row = APPEND_START_ROW + item_count - 1 if item_count > 0 else DATA_START_ROW
    ws[f"F{GROSS_VALUE_ROW}"].value = f"=SUM(F{first_row}:F{last_row})"
