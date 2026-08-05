"""Verification script for sheet-matching improvements."""
import sys

PASS = "OK  "
FAIL = "FAIL"
results = []


def check(label, condition):
    status = PASS if condition else FAIL
    results.append((status, label))
    print(f"{status}: {label}")


# ── 1. workbook_utils module exists and is importable ──────────────────────────
try:
    from excel_utils.workbook_utils import find_sheet, validate_workbook_sheets
    check("excel_utils.workbook_utils importable", True)
except ImportError as e:
    check(f"excel_utils.workbook_utils importable: {e}", False)
    sys.exit(1)

# ── 2. find_sheet: trailing space on real name ──────────────────────────────────
from openpyxl import Workbook

wb = Workbook(); wb.active.title = "JOB ORDER "
found = find_sheet(wb, "JOB ORDER")
check("find_sheet tolerates trailing space in real name", found.title == "JOB ORDER ")

# ── 3. find_sheet: leading space on real name ───────────────────────────────────
wb2 = Workbook(); wb2.active.title = "  SOE  "
found2 = find_sheet(wb2, "SOE")
check("find_sheet tolerates leading/trailing spaces in real name", found2.title == "  SOE  ")

# ── 4. find_sheet: case-insensitive ────────────────────────────────────────────
wb3 = Workbook(); wb3.active.title = "Proforma"
found3 = find_sheet(wb3, "proforma")
check("find_sheet is case-insensitive", found3.title == "Proforma")

# ── 5. find_sheet: diagnostic error message ─────────────────────────────────────
try:
    wb4 = Workbook(); wb4.active.title = "Other"
    find_sheet(wb4, "JOB ORDER")
    check("find_sheet raises ValueError for missing sheet", False)
except ValueError as e:
    msg = str(e)
    check("find_sheet raises ValueError for missing sheet", True)
    check("Error message includes normalized form", "normalized" in msg)
    check("Error message lists available sheets (raw)", "Available sheets (raw)" in msg)
    check("Error message lists available sheets (normalized)", "Available sheets (normalized)" in msg)

# ── 6. No old local definitions remain in writers ──────────────────────────────
import re

writer_files = [
    "excel_job_offer/writer.py",
    "excel_soe/writer.py",
    "excel_proforma/writer.py",
]
old_defs = [
    "def _get_job_order_sheet",
    "def _find_sheet",
    "def _get_sheet",
]
for fpath in writer_files:
    with open(fpath, encoding="utf-8") as f:
        content = f.read()
    for old in old_defs:
        check(
            f"{fpath}: no definition of '{old}'",
            old not in content,
        )

# ── 7. find_sheet imported in every writer ─────────────────────────────────────
for fpath in writer_files:
    with open(fpath, encoding="utf-8") as f:
        content = f.read()
    check(
        f"{fpath}: imports find_sheet from excel_utils",
        "from excel_utils.workbook_utils import find_sheet" in content,
    )

# ── 8. strip() applied to sheet constants ──────────────────────────────────────
strip_checks = [
    ("excel_job_offer/writer.py", '"JOB ORDER".strip()'),
    ("excel_soe/writer.py",       '"SOE".strip()'),
    ("excel_proforma/writer.py",  '"Proforma".strip()'),
]
for fpath, pattern in strip_checks:
    with open(fpath, encoding="utf-8") as f:
        content = f.read()
    check(f"{fpath}: constant uses .strip()", pattern in content)

# ── 9. config.py has REQUIRED_SHEETS and validate_template ─────────────────────
with open("app/config.py", encoding="utf-8") as f:
    cfg = f.read()
check("config.py defines REQUIRED_SHEETS", "REQUIRED_SHEETS" in cfg)
check("config.py defines validate_template()", "def validate_template" in cfg)

# ── 10. AGENTS.md mentions excel_utils ─────────────────────────────────────────
with open("AGENTS.md", encoding="utf-8") as f:
    agents = f.read()
check("AGENTS.md documents excel_utils", "excel_utils" in agents)
check("AGENTS.md documents find_sheet", "find_sheet" in agents)
check("AGENTS.md documents validate_workbook_sheets", "validate_workbook_sheets" in agents)

# ── Summary ────────────────────────────────────────────────────────────────────
passed = sum(1 for s, _ in results if s == PASS)
failed = sum(1 for s, _ in results if s == FAIL)
print()
print(f"{'='*50}")
print(f"TOTAL: {passed} passed, {failed} failed")
if failed:
    sys.exit(1)
