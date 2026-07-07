"""Write extracted job-over procedure lines to Excel."""

from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

JOB_ORDER_SHEET = "JOB ORDER "
TEXT_COLUMN = 3  # column C
NUMBER_COLUMN = 1  # column A
TEXT_MERGE_END_COLUMN = 12  # column L
ROW_BACKGROUND_LAST_COL = 22  # column V — full JOB ORDER procedure width
STYLE_COLUMNS = tuple(range(1, 7))
DEFAULT_TEMPLATE = Path("data.xlsm")
_NUMBERED_LINE = re.compile(r"^(\d+(?:\.\d+)*)\s+(.+)$")
_NO_FILL = PatternFill()
_EMPTY_SIDE = Side()
_WHITE_BACKGROUND_REFERENCE_ROW = 108
_BOX_RIGHT_REFERENCE_ROW = 33
_BOX_BOTTOM_REFERENCE_ROW = 34
_STYLE_TEMPLATE_ROWS = {
    "step": 100,
    "bullet": 89,
    "section": 81,
    "intro": 81,
    "note": 92,
    "text": 81,
}
_SECTION_FONT_ROW = 75
_DEFAULT_ROW_HEIGHT = 15.0


def write_job_offer_table(
    input_path: str | Path,
    output_path: str | Path,
    data: dict[str, Any],
) -> tuple[Path, int]:
    """Append procedure lines to the end of the JOB ORDER sheet."""
    input_path = Path(input_path)
    output_path = Path(output_path)
    lines = data.get("lines", [])
    if not lines:
        return output_path, 0

    keep_vba = input_path.suffix.lower() == ".xlsm"
    wb = load_workbook(input_path, keep_vba=keep_vba)
    ws = _get_job_order_sheet(wb)
    source = str(data.get("source", ""))

    write_row = _find_next_row(ws)
    previous_end_row = write_row - 1
    ws.insert_rows(write_row, amount=len(lines))

    for index, line in enumerate(lines):
        row = write_row + index
        text = str(line.get("text", "")).strip()
        kind, col_a, col_c = _prepare_row_values(text, source)
        _write_formatted_row(ws, row, kind, col_a, col_c)

    last_row = write_row + len(lines) - 1
    _extend_procedure_box(ws, write_row, last_row, previous_end_row)

    wb.save(output_path)
    return output_path, len(lines)


def _get_job_order_sheet(wb) -> Worksheet:
    if JOB_ORDER_SHEET not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet {JOB_ORDER_SHEET!r} not found in workbook. Available sheets: {available}"
        )
    return wb[JOB_ORDER_SHEET]


def _prepare_row_values(text: str, source: str) -> tuple[str, Any, str]:
    if text.startswith("Running Completion") and "DS," not in text:
        return "section", None, _format_text(text, source)

    if text.startswith("DS,"):
        return "intro", None, _format_text(text, source)

    if text.startswith(("!", "~")):
        body = text.lstrip("!~ ").strip()
        return "note", None, _format_text(f"NOTE: {body}", source)

    if text.startswith(("*", "-", "•", "●", "o ")):
        return "bullet", None, _format_bullet(text, source)

    step_number, body = _split_numbered_line(text)
    if step_number is not None:
        return "step", step_number, _format_text(body, source)

    return "text", None, _format_text(text, source)


def _format_text(text: str, source: str) -> str:
    if source in ("job_over_1c", "running_completion_summary"):
        return text.upper()
    return text


def _format_bullet(text: str, source: str) -> str:
    body = re.sub(r"^[\*\-•●o\s]+", "", text).strip()
    if source == "running_completion":
        return f"- {body}"
    return f"● {_format_text(body, source)}"


def _split_numbered_line(text: str) -> tuple[str | None, str]:
    if text.startswith(("*", "!", "~", "-", "•", "●", "o ")):
        return None, text

    match = _NUMBERED_LINE.match(text)
    if not match:
        return None, text

    number, body = match.group(1), match.group(2).strip()
    if "." in number:
        return number, body
    if body.lower().startswith("running"):
        return None, text
    return number, body


def _write_formatted_row(
    ws: Worksheet,
    row: int,
    kind: str,
    col_a: Any,
    col_c: str,
) -> None:
    template_row = _STYLE_TEMPLATE_ROWS.get(kind, _STYLE_TEMPLATE_ROWS["text"])
    _clear_row_merges(ws, row)
    _apply_row_background(ws, row)
    _copy_row_layout(ws, template_row, row)

    ws.cell(row, NUMBER_COLUMN).value = col_a
    ws.cell(row, TEXT_COLUMN).value = col_c

    if kind == "section":
        text_cell = ws.cell(row, TEXT_COLUMN)
        if not isinstance(text_cell, MergedCell):
            section_font = ws.cell(_SECTION_FONT_ROW, TEXT_COLUMN).font
            text_cell.font = Font(
                name=section_font.name or "Abadi",
                size=12,
                bold=True,
                color=copy(section_font.color),
            )

    _apply_row_background(ws, row)


def _copy_row_layout(ws: Worksheet, source_row: int, target_row: int) -> None:
    source_height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].height = (
        source_height if source_height and source_height <= 30 else _DEFAULT_ROW_HEIGHT
    )

    ws.merge_cells(
        start_row=target_row,
        start_column=TEXT_COLUMN,
        end_row=target_row,
        end_column=TEXT_MERGE_END_COLUMN,
    )

    for col in STYLE_COLUMNS:
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(source, MergedCell) or isinstance(target, MergedCell):
            source_col = _merged_top_left_col(ws, source_row, col)
            if source_col != col:
                continue
            source = ws.cell(source_row, source_col)
            target = ws.cell(target_row, source_col)
        if isinstance(target, MergedCell):
            continue
        target.number_format = copy(source.number_format)
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)


def _apply_row_background(ws: Worksheet, row: int) -> None:
    """Match the white/no-fill background used by rows above the gray template zone."""
    for col in range(1, ROW_BACKGROUND_LAST_COL + 1):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        reference = _background_reference_cell(ws, _WHITE_BACKGROUND_REFERENCE_ROW, col)
        cell.fill = copy(_NO_FILL if reference.fill.patternType else reference.fill)


def _get_procedure_box_border_sides(ws: Worksheet) -> tuple[Any, Any, Any]:
    left = copy(ws.cell(_WHITE_BACKGROUND_REFERENCE_ROW, NUMBER_COLUMN).border.left)
    right = copy(ws.cell(_BOX_RIGHT_REFERENCE_ROW, TEXT_MERGE_END_COLUMN).border.right)
    bottom = copy(ws.cell(_BOX_BOTTOM_REFERENCE_ROW, TEXT_MERGE_END_COLUMN).border.bottom)
    return left, right, bottom


def _set_border_side(cell, side: str, value) -> None:
    border = copy(cell.border)
    setattr(border, side, copy(value))
    cell.border = border


def _extend_procedure_box(
    ws: Worksheet,
    first_appended_row: int,
    last_row: int,
    previous_end_row: int,
) -> None:
    """Extend the blue procedure box to cover newly appended rows only."""
    if first_appended_row > last_row:
        return

    left_side, right_side, bottom_side = _get_procedure_box_border_sides(ws)

    if last_row > previous_end_row:
        _clear_row_bottom_border(ws, previous_end_row)

    for row in range(first_appended_row, last_row + 1):
        number_cell = ws.cell(row, NUMBER_COLUMN)
        if not isinstance(number_cell, MergedCell):
            _set_border_side(number_cell, "left", left_side)

        right_cell = _procedure_box_right_cell(ws, row)
        if not isinstance(right_cell, MergedCell):
            _set_border_side(right_cell, "right", right_side)

    _apply_row_bottom_border(ws, last_row, bottom_side)


def _procedure_box_right_cell(ws: Worksheet, row: int):
    """Return the cell that owns the right edge of the procedure text area."""
    cell = ws.cell(row, TEXT_MERGE_END_COLUMN)
    if isinstance(cell, MergedCell):
        top_left_col = _merged_top_left_col(ws, row, TEXT_MERGE_END_COLUMN)
        return ws.cell(row, top_left_col)
    return cell


def _clear_row_bottom_border(ws: Worksheet, row: int) -> None:
    for col in range(NUMBER_COLUMN, TEXT_MERGE_END_COLUMN + 1):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        _set_border_side(cell, "bottom", _EMPTY_SIDE)


def _apply_row_bottom_border(ws: Worksheet, row: int, bottom_side) -> None:
    for col in range(NUMBER_COLUMN, TEXT_MERGE_END_COLUMN + 1):
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            continue
        _set_border_side(cell, "bottom", bottom_side)


def _clear_row_merges(ws: Worksheet, row: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row:
            try:
                ws.unmerge_cells(str(merged))
            except KeyError:
                pass


def _merged_top_left_col(ws: Worksheet, row: int, col: int) -> int:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return col
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return merged.min_col
    return col


def _background_reference_cell(ws: Worksheet, row: int, col: int):
    ref_col = _merged_top_left_col(ws, row, col)
    return ws.cell(row, ref_col)


def _find_next_row(ws: Worksheet) -> int:
    last_row = 1
    for row in range(1, ws.max_row + 1):
        if any(_cell_value(ws, row, col) for col in range(1, 7)):
            last_row = row
    return last_row + 1


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(merged.min_row, merged.min_col).value
        return None
    return cell.value
